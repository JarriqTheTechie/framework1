import mysql.connector
from argparse import Action
from framework1.core_services.Request import Request
from framework1.database.ActiveRecord import PaginationResult, ActiveRecord
from framework1.dsl.TableFilter import Filter
from framework1.service_container._Injector import injectable_route
from framework1.utilities.DataKlass import DataKlass
from markupsafe import Markup, escape
from typing import Self, Callable, Union
from framework1.cli.migrate import discover_models
from framework1 import render_template_string_safe_internal
from framework1.database.QueryBuilder import QueryBuilder
from dataclasses import dataclass, asdict
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import send_file

from app import app

@injectable_route(app, '/f1/export-excel', methods=['GET'])
def TableExportExcel(request: Request):
    model_class_name = request.input("model")
    if not model_class_name:
        return {"success": False, "message": "Model required"}

    model_map = {
        "ActivityTable": Activity,
        # add other mappings as needed
    }

    model_cls = model_map.get(model_class_name)
    if not model_cls:
        return {"success": False, "message": "Unknown model"}

    table = model_cls()  # assume table classes extend Table
    table.paginate(page=1, per_page=1000000)  # or fetch all
    return table.export_excel(filename=f"{model_class_name}.xlsx")

@injectable_route(app, '/f1/delete-bulk', methods=['GET'])
def TableDeleteBulk(request: Request):
    # TODO: WORK ON BULK ACTIONS
    models = discover_models("lib")
    ids = request.to_list("ids", cast=int)
    model_class_name = request.input("model")
    if not ids or not model_class_name:
        return {"success": False, "message": "Invalid request"}

    model_map = {
        "ActivityTable": Activity,
        # Add other mappings as needed
    }

    model_cls = model_map.get(model_class_name)
    if not model_cls:
        return {"success": False, "message": "Unknown model"}

    model_cls().where_in("id", ids).delete()
    return {"success": True}


def record_to_dict(record) -> dict:
    if hasattr(record, "to_dict") and not isinstance(record, DataKlass):
        return record.to_dict()
    elif hasattr(record, "to_dict") and isinstance(record, DataKlass):
        return record
    elif isinstance(record, ActiveRecord):
        return record
    else:
        return record


class Field:
    def __init__(self, name):
        """The only required argument is the name."""
        self.__name = name
        self.__header = name  # Default to the name if no header is set
        self.__class_name = ""
        self.__modify_using_ = None
        self.__value_if_missing = ""
        self._icon_color = ""
        self._icon = None
        self._icon_position = "left"  # Default icon position
        self._icon_map = []
        self._outer_class = ""
        self._badge = False
        self._badge_color_map = []
        self._static_badge_color = None
        self._date_format = None
        self._render_html = False
        self._description = None
        self._description_position = 'below'  # or 'above'
        self._limit = None
        self._limit_end = '...'
        self._words_limit = None
        self._words_end = '...'
        self._sortable = False
        self._searchable = False
        self._tooltip = None
        self._extra_attributes = None
        self._extra_cell_attributes = None
        self._hidden = False

    def name(self):
        return self.__name

    def header(self):
        return self.__header

    def class_name(self):
        return self.__class_name

    def label(self, header: str):
        self.__header = header
        return self

    def placeholder(self, placeholder: str):
        self.__value_if_missing = placeholder
        return self

    def classes(self, class_name: str):
        self.__class_name = f"{self.__class_name} {class_name}" if self.__class_name else class_name
        return self

    def color(self, color: str):
        """Set the color class for the field."""
        if color:
            self.__class_name = f"{self.__class_name} text-{color}" if self.__class_name else f"text-{color}"
        return self

    def icon(self, icon: str | dict[str, str]):
        if isinstance(icon, str):
            self._icon = icon
        else:
            if isinstance(icon, dict):
                self._icon_map = icon
            else:
                raise TypeError("Icon must be a string or a dictionary mapping values to icons")
        return self

    def icon_position(self, position: str):
        """Set the position of the icon."""
        if position in ["left", "right"]:
            self._icon_position = position
        else:
            raise ValueError("Icon position must be 'left' or 'right'")
        return self

    def icon_color(self, color: str):
        """Set the color of the icon."""
        if color:
            self._icon_color = f"text-{color}"
        return self

    def badge(self):
        """Add a badge to the field."""
        self._badge = True
        return self

    def badge_color(self, color: str | dict[str, str]):
        """
        Set badge color either statically or using a value mapping.

        Args:
            color: Either a string for static color or dict mapping values to colors
        """
        if isinstance(color, str):
            self._static_badge_color = color
        else:
            self._badge_color_map = color
        return self

    def modify_using(self, modify_using_: callable):
        self.__modify_using_ = modify_using_
        return self

    def default(self, value_if_missing: str):
        self.__value_if_missing = value_if_missing
        return self

    def _format_value(self, value, record):
        """Apply modifier if exists, otherwise return the value."""
        if value is None or value == "":
            value = self.__value_if_missing
        if self.__modify_using_:
            try:
                return self.__modify_using_(value, record)
            # except TypeError:
            #     return self.__modify_using_(value)
            except Exception as e:
                raise e
        return value

    def date(self, format_string: str = "%Y-%m-%d %H:%M:%S"):

        def date_formatter(value, _):
            if not value:
                return ""

            python_format = format_string

            try:
                from datetime import datetime
                if isinstance(value, str):
                    # Try parsing the string as datetime
                    try:
                        # Try ISO format first
                        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                    except ValueError:
                        # Try common formats
                        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                            try:
                                dt = datetime.strptime(value, fmt)
                                break
                            except ValueError:
                                continue
                        else:
                            return value  # Return original if parsing fails
                elif isinstance(value, (int, float)):
                    # Assume timestamp
                    dt = datetime.fromtimestamp(value)
                else:
                    # Assume it's already a datetime object
                    dt = value

                return dt.strftime(python_format)
            except Exception:
                return value  # Return original value if formatting fails

        self.modify_using(date_formatter)
        return self

    def html(self):
        """
        Allow the field value to be rendered as HTML.

        Returns:
            Self: Returns the field instance for method chaining
        """
        self._render_html = True
        return self

    def description(self, description_value: str | Callable, position: str = 'below', limit: int = 100,
                    end: str = '...', html=False):
        """
        Add a description to the field content.
        
        Args:
            description_value: String or callable that takes the record and returns a string
            position: Position of description - 'below' or 'above' (default: 'below')
        """
        if position not in ['below', 'above']:
            raise ValueError("Position must be either 'below' or 'above'")

        self._description = description_value
        self._description_position = position
        self._description_limit = limit
        self._description_limit_end = end
        self._description_is_html = html
        return self

    def limit(self, count: int | Callable, end: str = '...'):
        """
        Limit the length of the column's value.
        
        Args:
            count: Integer or callable that returns the maximum length
            end: String to append when text is truncated (default: '...')
        """
        self._limit = count
        self._limit_end = end
        return self

    def words(self, count: Union[int, Callable], end: str = '...'):
        """
        Limit the number of words in the column's value.
        """
        self._words_limit = count
        self._words_end = end
        return self

    def sortable(self):
        self._sortable = True
        return self

    def searchable(self):
        self._searchable = True
        return self

    def url(self, url_pattern):
        """
        Set a URL pattern or function for the field.
        Args:
            url_pattern (str | callable): URL template with placeholders (e.g., "/users/{id}")
                                          or a callable (record -> url)
        """
        self._url_template = url_pattern
        return self

    def tooltip(self, text_or_callable):
        """
        Set a tooltip for this field. Can be:
        - A static string
        - A callable that takes (record) or (record, data) and returns a string
        """
        self._tooltip = text_or_callable
        return self

    def extra_attributes(self, attrs):
        """
        Set additional HTML attributes for the <td> tag.
        Example: {'class': 'slug-column', 'data-slug': 'value'}
        """
        self._extra_attributes = attrs
        return self

    def extra_cell_attributes(self, attrs):
        """
        Set additional attributes for the <td> tag (cell-specific).
        Supports static dict or callable(record).
        """
        self._extra_cell_attributes = attrs
        return self

    def hidden(self, value: bool | Callable = True):
        """
        Mark the field as hidden.
        Args:
            value (bool | Callable): True/False or a callable(record) -> bool
        """
        self._hidden = value
        return self


class TextColumn(Field):
    @staticmethod
    def make(name):
        return TextColumn(name)


class Table:
    table_class = ""
    table_style = ""
    thead_class = ""  # Class for <thead>
    tbody_class = ""  # Class for <tbody>
    tr_class = ""  # Class for <tr>

    key_id = "id"
    search_key = "id"
    model = None

    persist_sort = False
    persist_search = False
    persist_filters = False
    selectable = False
    master_detail_expandable = True

    filterable_fields = []

    search_placeholder = "Search..."

    def __init__(self, data: list[dict] = None, non_activerecord_model=None, as_dto=False, dto_target=None):
        self.pagination = None
        self.data = data
        self.query = None
        self.sub_resource_table = False
        self.table_name = self.__class__.__name__

        # If a model is set, initialize its query builder
        try:
            self.query = self.model() if self.model else None
        except TypeError:
            self.query = self.model if self.model else None

        # If user overrides modify_table_query, apply it
        if self.modify_table_query.__func__ is not Table.modify_table_query:
            # Step 1: Get normalized filters from the form
            filters = Request().grouped("filters")
            grouped_filters = {}

            # Step 2: Group filters by 'group'
            for f in filters:
                group_name = f.get('group', 'default') or 'default'
                grouped_filters.setdefault(group_name, []).append(f)

            # Step 3: Apply each group of filters to the query
            for _, group_filters in grouped_filters.items():
                self.query = self.apply_filter_conditions(self.query, group_filters)
            self.query = self.modify_table_query()
        else:
            # Step 1: Get normalized filters from the form
            filters = Request().grouped("filters")
            grouped_filters = {}

            # Step 2: Group filters by 'group'
            for f in filters:
                group_name = f.get('group', 'default') or 'default'
                grouped_filters.setdefault(group_name, []).append(f)

            # Step 3: Apply each group of filters to the query
            for _, group_filters in grouped_filters.items():
                self.query = self.apply_filter_conditions(self.query, group_filters)

        # If a non-ActiveRecord model is passed, use it directly
        if non_activerecord_model:
            if hasattr(non_activerecord_model, 'paginate'):
                self.query = non_activerecord_model
            else:
                raise TypeError("non_activerecord_model must be an instance of QueryBuilder")

        # If we have a query, apply automatic sorting
        if self.query:
            self.query = self._apply_search(self.query)
            self.query = self._apply_sorting(self.query)

        if hasattr(self, "filters"):
            self.query = self._apply_filters_grouped(self.filters(), self.query)

        # If user supplied static data
        if self.data is None and self.query is not None:
            # Load all records (if no pagination is used yet)
            self.data = self.query.get()

        # If user wants DTOs
        if as_dto and self.model and dto_target:
            self.as_dto = as_dto
            self.dto_target = dto_target

    def apply_filter_conditions(self, query, filters: list[dict]):
        """
        Applies a list of filter dictionaries to the query, handling AND/OR and nesting.
        """

        def apply_group(q):
            for idx, cond in enumerate(filters):
                field = cond["field"]
                op = cond["operator"]
                val = cond.get("value", "")
                boolean = cond.get("boolean", "and").lower()
                values = [v.strip() for v in val.split(',')] if val else []

                # Determine AND vs OR application
                method = q.where if boolean == "and" or idx == 0 else q.or_where

                # Map operators to QueryBuilder methods
                match op:
                    case "where":
                        q = method(field, val)
                    case "not_equal":
                        q = method(field, '!=', val)
                    case "contains":
                        q = method(field, 'LIKE', f"%{val}%")
                    case "starts_with":
                        q = method(field, 'LIKE', f"{val}%")
                    case "ends_with":
                        q = method(field, 'LIKE', f"%{val}")
                    case "greater_than":
                        q = method(field, '>', val)
                    case "less_than":
                        q = method(field, '<', val)
                    case "greater_than_eq":
                        q = method(field, '>=', val)
                    case "less_than_eq":
                        q = method(field, '<=', val)
                    case "in" if values:
                        q = method(field, 'IN', values)
                    case "not_in" if values:
                        q = method(field, 'NOT IN', values)
                    case "between" if len(values) == 2:
                        # Support both where_between_dates and or_where_between_dates
                        between_method_name = f"{method.__name__}_between_dates"
                        if hasattr(q, between_method_name):
                            q = getattr(q, between_method_name)(field, *values)
                    case "is_null":
                        q = method(field, 'IS', None)
                    case "is_not_null":
                        q = method(field, 'IS NOT', None)
                    case "regex":
                        q = method(field, 'REGEXP', val)

            return q

        # If multiple filters in the same group, nest them for proper grouping
        return query.nest(apply_group) if len(filters) > 1 and hasattr(query, 'nest') else apply_group(query)

    def _apply_filters_grouped(self, filters: list[Filter], query: QueryBuilder) -> QueryBuilder:
        request = Request()
        session = request.session()

        grouped: dict[str, list[Filter]] = {}
        ungrouped: list[Filter] = []

        # 1. Separate grouped and ungrouped filters
        for f in filters:
            group_key = getattr(f, "_group_key", None)
            if group_key:
                grouped.setdefault(group_key, []).append(f)
            else:
                ungrouped.append(f)

        # 2. Apply grouped filters using or_where
        for group_filters in grouped.values():
            for i, f in enumerate(group_filters):
                value = request.input(f"filter_" + f._key)
                if self.persist_filters:
                    value = value or session.get(f"{f._key}_filter")

                if value or (value is None and getattr(f, "_default_checked", False)):
                    if f._query_callback:
                        if i == 0:
                            query = f._query_callback(query)
                        else:
                            # ✅ Assume your callback is always: lambda q: q.where(field, op, val)
                            # → We must extract those args manually
                            # So instead of using the callback again, hard-code the clause
                            # OR force a known callback format
                            # But since we're not going fancy:
                            # We reapply manually using or_where directly
                            # This requires you to know the field/val here (see below workaround)

                            # Instead of reusing the callback (which uses `where`)
                            # You’ll need to pass field/operator/value at filter setup
                            pass  # we’ll improve this below 👇

        # 3. Apply ungrouped filters
        for f in ungrouped:
            query = f.apply(query, self.persist_filters)

        return query

    def paginate(self, page: int = None, per_page: int = None):
        """Use database-level pagination."""
        if not self.model:
            raise ValueError("Model class must be set to use pagination")



        request = Request()
        table_name = self.__class__.__name__
        instance_table = request.grouped(table_name)
        print(instance_table)

        if instance_table:
            page = page or int(instance_table.get(table_name).get('page', 1))
            per_page = per_page or int(instance_table.get(table_name).get('per_page', 10))
        else:
            page = page or request.integer('page', 1)
            per_page = per_page or request.integer('per_page', 10)


        # Use the query builder from the model for pagination
        if hasattr(self.query.__class__, '__primary_key__') or getattr(self.query.__class__, '__driver__') == "mssql":
            if not self.query.order_by_clauses:
                self.pagination = self.query.order_by(self.query.__class__.__primary_key__, "asc").paginate(page,
                                                                                                            per_page)
            else:
                self.pagination = self.query.paginate(page, per_page)
        else:

            self.pagination = self.query.paginate(page, per_page)

        if not isinstance(self.pagination, PaginationResult):
            self.pagination = {}
            self.data = []
            return self
        if not getattr(self, "dto_target", None):
            self.data = self.pagination.items
        else:
            self.data = self.pagination.items if not self.as_dto else self.pagination.items.to_dtos(self.dto_target)
        return self

    def export_excel(self, filename: str = None):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        import io
        from flask import send_file

        fields = [
            f for f in self.schema()
            if hasattr(f, "_hidden") and not (callable(f._hidden) and f._hidden({}))
               and not (isinstance(f._hidden, bool) and f._hidden)
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = self.__class__.__name__

        # --- Custom header hook ---
        if hasattr(self, "get_export_header"):
            header_text = self.get_export_header()
            if header_text:
                ws.append([header_text])
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
                cell = ws.cell(row=1, column=1)
                cell.font = Font(size=14, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.append([])  # blank row

        # --- Column headers ---
        headers = [f.header() for f in fields]
        ws.append(headers)
        header_row_index = ws.max_row
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=header_row_index, column=col)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        # --- Rows ---

        for record in self.data:
            record_dict = record_to_dict(record)
            ws.append([field._format_value(record_dict.get(field.name(), ""), record_dict) for field in fields])

        # --- Custom footer hook ---
        if hasattr(self, "get_export_footer"):
            footer_text = self.get_export_footer()
            if footer_text:
                ws.append([])
                footer_row = ws.max_row + 1
                ws.cell(row=footer_row, column=1, value=footer_text)
                ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(fields))
                cell = ws.cell(row=footer_row, column=1)
                cell.font = Font(italic=True, color="FF666666")
                cell.alignment = Alignment(horizontal="right")

        # --- Auto-fit columns ---
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        download_name = filename or f"{self.__class__.__name__}.xlsx"
        return send_file(output,
                         as_attachment=True,
                         download_name=download_name,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def schema(self) -> list[Field]:
        """Override this method to define schema."""
        return []

    def set_as_sub_resource_table(self):
        self.sub_resource_table = True
        return self

    def set_key_id(self, key_id: str):
        self.key_id = key_id
        return self

    def has_default_actions(self) -> bool:
        """Override this method to disable default actions column."""
        return True

    def has_custom_actions(self) -> bool:
        """Override this method to enable custom actions column."""
        return False

    def get_custom_actions(self, record) -> list[Action]:
        """Override this method to provide custom actions."""
        return []

    def render(self) -> Markup:
        """Generate HTML for the table with improved configurability."""
        request = Request()
        fields = self.schema()

        def build_table_header(fields: list[Field]) -> list[str]:
            header = [f'<thead class="{self.thead_class}"><tr>']
            if getattr(self, 'selectable'):
                header.append('<th class="text-center"><input type="checkbox" class="select-all"></th>')
            existing_fields = request.input(f"{self.table_name}[sort]", "").split(",")
            existing_dirs = request.input(f"{self.table_name}[sort_dir]", "").split(",")

            for field in fields:
                if isinstance(field, MasterDetailRow):
                    continue

                is_hidden = field._hidden if not callable(
                    field._hidden) else False  # headers don't use per-row callable
                if is_hidden:
                    continue

                th_classes = field.class_name()
                content = field.header()

                if getattr(field, "_sortable", False):
                    if field.name() in existing_fields:
                        idx = existing_fields.index(field.name())
                        current_dir = existing_dirs[idx] if idx < len(existing_dirs) else "asc"
                        next_dir = "desc" if current_dir == "asc" else "asc"
                    else:
                        next_dir = "asc"

                    new_fields = existing_fields.copy()
                    new_dirs = existing_dirs.copy()

                    if field.name() in existing_fields:
                        idx = existing_fields.index(field.name())
                        new_dirs[idx] = next_dir
                    else:
                        new_fields.append(field.name())
                        new_dirs.append(next_dir)

                    query_args = request.all()
                    query_args[f"{self.table_name}[sort]"] = ",".join(new_fields)
                    query_args[f"{self.table_name}[sort_dir]"] = ",".join(new_dirs)

                    from urllib.parse import urlencode
                    sort_url = f"{request.path()}?{urlencode(query_args)}"

                    icon = ""
                    if field.name() in existing_fields:
                        idx = existing_fields.index(field.name())
                        dir_icon = existing_dirs[idx] if idx < len(existing_dirs) else "asc"
                        icon = f' <i class="ri-arrow-{"up-long-line" if dir_icon == "asc" else "down-long-line"}"></i>'

                    content = f'<a href="{sort_url}">{content}{icon}</a>'

                header.append(f'<th class="{th_classes}">{content}</th>')
            # TODO: SUPPORT TABLE ROW ACTIONS
            # header.append('<th class="">Actions</th>')
            header.append('</tr></thead>')
            return header

        def build_cell_content(field: Field, value: str, record) -> str:
            if isinstance(field, MasterDetailRow):
                formatted_value = value
            else:
                formatted_value = field._format_value(value, record)

            record = record_to_dict(record)

            # Handle character limit
            if hasattr(field, '_limit') and field._limit is not None:
                limit_value = field._limit(record) if callable(field._limit) else field._limit
                if isinstance(limit_value, int) and len(str(formatted_value)) > limit_value:
                    formatted_value = str(formatted_value)[:limit_value] + field._limit_end

            # Handle word limit
            if hasattr(field, '_words_limit') and field._words_limit is not None:
                words_limit = field._words_limit(record) if callable(field._words_limit) else field._words_limit
                if isinstance(words_limit, int):
                    words = str(formatted_value).split()
                    if len(words) > words_limit:
                        formatted_value = ' '.join(words[:words_limit]) + field._words_end

            # Handle HTML escaping
            if not getattr(field, '_render_html', False):
                from markupsafe import escape
                formatted_value = escape(formatted_value)

            content_parts = []

            # Handle icon
            icon_classes = []
            if getattr(field, '_icon_map', None) and formatted_value in field._icon_map:
                icon_classes.append(field._icon_map[formatted_value])
            if getattr(field, '_icon', None):
                icon_classes.append(self._icon)
            if getattr(field, '_icon_color', None):
                icon_classes.append(field._icon_color)
            icon_html = f'<i class="{" ".join(icon_classes)}"></i>'
            if getattr(field, '_icon_position', 'left') == 'left':
                content_parts.append(f'{icon_html} ')
            else:
                content_parts.append(f' {icon_html}')

            # Add the main content
            content_parts.insert(
                1 if getattr(field, '_icon_position', 'left') == 'left' else 0,
                formatted_value
            )

            content = ''.join(content_parts)

            # Add description if present
            if getattr(field, '_description', None):
                if callable(field._description):
                    from inspect import signature
                    sig = signature(field._description)
                    if len(sig.parameters) == 2:
                        description_text = field._description(record, record)
                    else:
                        description_text = field._description(record)
                else:
                    description_text = field._description

                if description_text:
                    if not field._description_is_html:
                        content += f'<div class="text-muted small">{escape(str(description_text)[:field._description_limit])}{field._description_limit_end}</div>'
                    else:
                        content += f'<div class="text-muted small">{Markup(description_text[:field._description_limit])}{field._description_limit_end}</div>'

            # Handle badge
            if getattr(field, '_badge', False):
                badge_classes = ['badge']
                if getattr(field, '_badge_color_map', None) and formatted_value in field._badge_color_map:
                    badge_classes.append(f'bg-{field._badge_color_map[formatted_value]}')
                elif getattr(field, '_static_badge_color', None):
                    badge_classes.append(f'bg-{field._static_badge_color}')
                content = f'<span class="{" ".join(badge_classes)}">{content}</span>'

            # Handle URL
            if getattr(field, '_url_template', None):
                if callable(field._url_template):
                    url = field._url_template(record)
                else:
                    try:
                        url = field._url_template.format(record)
                    except KeyError:
                        url = "#"
                content = f'<a href="{url}">{content}</a>'

            tooltip_html_open = tooltip_html_close = ""
            tooltip_text = None

            if getattr(field, '_tooltip', None):
                tooltip = field._tooltip
                if callable(tooltip):
                    from inspect import signature
                    sig = signature(tooltip)
                    if len(sig.parameters) == 2:
                        tooltip_text = tooltip(record, record)
                    else:
                        tooltip_text = tooltip(record)
                else:
                    tooltip_text = tooltip

            if tooltip_text:
                tooltip_html_open = f'<span data-bs-toggle="tooltip" title="{escape(tooltip_text)}">'
                tooltip_html_close = '</span>'

            content = f'{tooltip_html_open}{content}{tooltip_html_close}'

            return content

        def build_table_body(data: list, fields: list[Field]) -> list[str]:
            body = [f'<tbody class="{self.tbody_class}">']
            has_default_actions = self.has_default_actions()
            has_custom_actions = self.has_custom_actions()
            for record in data:
                record = record_to_dict(record)

                row_dbl_click_action_html = ""
                try:
                    row_dbl_click_action = self.record_url(record)
                    row_dbl_click_action_html = f' ondblclick="{row_dbl_click_action}"'
                except AttributeError:
                    pass

                row = [f'<tr class="{self.tr_class}" {row_dbl_click_action_html}>']
                if getattr(self, 'selectable'):
                    row.append(
                        f'<td class="text-center"><input type="checkbox" class="row-select" value="{record.get(self.key_id)}"></td>')

                for field in fields:
                    if isinstance(field, MasterDetailRow):
                        continue

                    is_hidden = False
                    if callable(field._hidden):
                        is_hidden = field._hidden(record)
                    else:
                        is_hidden = field._hidden

                    if is_hidden:
                        continue  # Skip this cell

                    if "." in field.name():
                        # Support nested fields like 'user.name'
                        parts = field.name().split(".")
                        value = record
                        for part in parts:
                            value = value.get(part, "")
                            if not value:
                                break
                    else:
                        value = record.get(field.name(), "")

                    content = build_cell_content(field, value, record)

                    # Build extra cell attributes
                    attr_dict = {}
                    if field._extra_cell_attributes:
                        if callable(field._extra_cell_attributes):
                            attr_dict = field._extra_cell_attributes(record)
                        else:
                            attr_dict = field._extra_cell_attributes

                    # Merge class names
                    base_class = field.class_name()
                    extra_class = attr_dict.pop('class', '')
                    combined_class = f"{base_class} {extra_class}".strip()

                    # Build other attribute string, escape values for safety
                    attr_str = ' '.join(f'{k}="{escape(v)}"' for k, v in attr_dict.items())
                    if field.name() == fields[0].name():
                        collapse_toggle = f'<button class="collapse-caret" type="button" data-bs-toggle="collapse" data-bs-target=".collapse-contentId-{record.get(self.key_id)}" aria-expanded="false" aria-controls="contentId-{record.get(self.key_id)}"> <i class="ri-arrow-right-s-line"></i> </button>'
                        row.append(
                            f'<td data-framework1-field-name="{field.name()}" class="{combined_class}" {attr_str}>{collapse_toggle} {content}</td>'
                        )
                    else:
                        row.append(
                            f'<td data-framework1-field-name="{field.name()}" class="{combined_class}" {attr_str}>{content}</td>'
                        )

                # TODO: SUPPORT TABLE ROW ACTIONS
                if has_default_actions and has_custom_actions:
                    row.append(
                        f'<td data-framework1-field-name="" class="">Delete {record_data}</td>'
                    )
                    row.append(self.get_custom_actions(record_data))

                row.append('</tr>')
                body.extend(row)

                # --- MASTER DETAIL SUPPORT ---
                if getattr(self, "master_detail_expandable", None):
                    # Rerun build_table_body but pass the full record as data
                    # nested_table_html = build_table_body([record], fields)
                    for row in fields:
                        if isinstance(row, MasterDetailRow):
                            if getattr(row, '_template', None):
                                if callable(row._template):
                                    master_detail_view_template = row._template(record)
                                else:
                                    master_detail_view_template = row._template
                                body.append(
                                    f'''
                                    <tr class="master-detail-row collapse collapse-contentId-{record.get(self.key_id)}">
                                        <td colspan="{len(fields) + (1 if getattr(self, 'selectable', False) else 0)}">
                                            <div class="collapse collapse-contentId-{record.get(self.key_id)}">
                                                {master_detail_view_template}
                                            </div>
                                        </td>
                                    </tr>
                                    '''
                                )
                                break

            body.append('</tbody>')
            return body

        def build_pagination() -> list[str]:
            if not (pagination := getattr(self, 'pagination', None)):
                return []

            if not hasattr(pagination, 'items'):
                return []

            total = pagination.total
            current_page = pagination.current_page
            last_page = pagination.last_page

            return [
                render_template_string_safe_internal("table-dsl/pagination.html",
                                                     total=total,
                                                     data=self.data,
                                                     current_page=current_page,
                                                     last_page=last_page,
                                                     per_page=pagination.per_page,
                                                     table_name=self.__class__.__name__,
                                                     request=request)
            ]

        # Main table assembly
        fields = self.schema()
        html = [
            f'<div class="table-responsive">',
            f'<table id="{self.__class__.__name__}" class="{self.table_class}" style="{self.table_style if getattr(self, "table_style", "") else ""}">'
        ]

        search_session_key = f"{self.__class__.__name__}_search"
        search_value = escape(Request().input("search", request.session().get(search_session_key, "")))
        search_placeholder = self.search_placeholder

        table_actions_header = []

        if not bool(self.sub_resource_table):
            table_actions_header.insert(0, render_template_string_safe_internal("table-dsl/search.html",
                                                                                search_value=search_value,
                                                                                search_placeholder=search_placeholder))

        html.insert(0,
                    f"<div class='table-actions d-inline-flex  mb-3 justify-content-end'>{''.join(table_actions_header)}</div>")

        html.extend(build_table_header(fields))

        if self.data:
            html.extend(build_table_body(self.data, fields))

        html.append('</table>')
        html.extend(build_pagination())
        html.append('</div>')

        from framework1.dsl.F1TableFilterForm import F1TableFilterForm

        if len(self.filterable_fields) != 0:
            filter_form = F1TableFilterForm(request.all()).set_resource_from_table(self)
            filter_bar_css = render_template_string_safe_internal('table-dsl/filter-bar-styles.html')
            filter_bar = render_template_string_safe_internal("table-dsl/filter-bar.html", filter_form=filter_form,
                                                              filter_bar_css=filter_bar_css)
            html.insert(1, filter_bar)

        return Markup('\n'.join(html))

    def __str__(self) -> Self:
        """Return HTML when the object is converted to a string."""
        return self.render()

    def _apply_sorting(self, query):
        request = Request()
        session_key = f"{self.__class__.__name__}_sort"

        sort_fields = request.input(f"{self.table_name}[sort]", "").split(",") if request.has(f"{self.table_name}[sort]") else []
        sort_dirs = request.input(f"{self.table_name}[sort_dir]", "").lower().split(",") if request.has(f"{self.table_name}[sort_dir]") else []

        valid_sort_fields = [f.name() for f in self.schema() if getattr(f, "_sortable", False)]
        applied_sort = False

        # Apply user-provided sort
        for idx, field in enumerate(sort_fields):
            field = field.strip()
            if field in valid_sort_fields:
                dir_ = sort_dirs[idx] if idx < len(sort_dirs) else "asc"
                dir_ = dir_ if dir_ in ["asc", "desc"] else "asc"
                query = query.order_by(field, dir_)
                applied_sort = True

        if applied_sort and self.persist_sort:
            # Save to session
            request.session()[session_key] = {
                f"{self.table_name}[sort]": ",".join(sort_fields),
                f"{self.table_name}[sort_dir]": ",".join(sort_dirs)
            }

        # If no sort applied, check session if persist_sort enabled
        if not applied_sort and self.persist_sort:
            session_sort = request.session().get(session_key)
            if session_sort:
                s_fields = session_sort.get(f"{self.table_name}[sort]", "").split(",")
                s_dirs = session_sort.get(f"{self.table_name}[sort_dir]", "").lower().split(",")
                for idx, field in enumerate(s_fields):
                    field = field.strip()
                    if field in valid_sort_fields:
                        dir_ = s_dirs[idx] if idx < len(s_dirs) else "asc"
                        dir_ = dir_ if dir_ in ["asc", "desc"] else "asc"
                        query = query.order_by(field, dir_)
                        applied_sort = True

        # Fallback to default sort
        if not applied_sort:
            default_field, default_dir = self.default_sort()
            if default_field and default_field in valid_sort_fields:
                default_dir = default_dir.lower() if default_dir else "asc"
                query = query.order_by(default_field, default_dir)

        return query

    def _apply_search(self, query):
        request = Request()
        session = request.session()
        session_key = f"{self.__class__.__name__}_search"

        search_term = request.input("search")

        if getattr(self, "persist_search", False):
            if search_term is not None:
                if search_term.strip() == "":
                    session.pop(session_key, None)
                    search_term = None
                else:
                    session[session_key] = search_term
            else:
                search_term = session.get(session_key)

        if not search_term:
            return query

        # Merge method + column searchables
        method_fields = []
        if hasattr(self, "searchable") and self.searchable.__func__ is not Table.searchable:
            method_fields = self.searchable()

        column_fields = [f.name() for f in self.schema() if getattr(f, "_searchable", False)]
        if isinstance(self.search_key, str):
            column_fields.append(self.search_key)
        elif isinstance(self.search_key, list):
            column_fields.extend(self.search_key)
        searchable_fields = list(dict.fromkeys(method_fields + column_fields))

        if not searchable_fields:
            return query

        # Split search into terms
        terms = [t.strip() for t in search_term.strip().split() if t.strip()]
        if not terms:
            return query

        # Apply where_any_columns for each term
        for term in terms:
            query = query.where_any_columns(searchable_fields, "LIKE", f"%{term}%")

        return query

    def modify_table_query(self):
        """Override this method to modify the query for the table."""
        pass

    def default_sort(self) -> tuple[str, str]:
        """
        Override this in your table subclass to provide a default sort.
        Example: return ("id", "asc")
        """
        return None, None

    def searchable(self) -> list[str]:
        """
        Override this in your table subclass to provide searchable fields.
        Example: return ["name", "email"]
        """
        return []

    def filters(self) -> list[Filter]:
        """
        Override this in your table subclass to provide filters.
        Example: return [Filter.make("active").label("Active").query(lambda q: q.where_active())]
        """
        return []


class MasterDetailRow:
    def __init__(self, name: str):
        self.name = name
        self.data = ""
        self.raw_html = ""
        self.fields = []

    @staticmethod
    def make(name):
        return MasterDetailRow(name)

    def schema(self, fields: list[Field] = None) -> list[Field]:
        self.fields = fields if fields else []
        return self

    def set_data(self, data):
        self.data = data
        return self

    def template(self, template: str | Callable) -> Self:
        self._template = template
        return self
