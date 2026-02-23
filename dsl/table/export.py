from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from flask import send_file
import io

from .utils import record_to_dict


class TableExportMixin:
    def export_excel(self, filename: str = None):
        fields = [
            f
            for f in self._get_schema_cached()
            if hasattr(f, "_hidden")
            and not (callable(f._hidden) and f._hidden({}))
            and not (isinstance(f._hidden, bool) and f._hidden)
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = self.__class__.__name__

        # --- Custom header hook ---
        if hasattr(self, "get_export_header"):
            header_text = self.get_export_header()
            if header_text:
                ws.append([header_text])
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
                cell = ws.cell(row=1, column=1)
                cell.font = Font(size=14, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.append([])  # blank row

        # --- Column headers ---
        headers = [f.header() for f in fields]
        ws.append(headers)
        header_row_index = ws.max_row
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=header_row_index, column=col)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        # --- Rows ---
        for record in self._ensure_data_loaded(force_query_get=True):
            record_dict = record_to_dict(record)

            # Guard: ensure we always have a mapping for field lookups
            if not hasattr(record_dict, "get"):
                if hasattr(record, "to_dict"):
                    record_dict = record.to_dict()
                elif isinstance(record_dict, dict):
                    pass
                else:
                    # Fallback to single-value mapping to avoid hard failure
                    record_dict = {"value": record_dict}

            row_values = []
            for field in fields:
                # Support nested fields like "client.Name" (mirrors render.py)
                if "." in field.name():
                    value = record_dict
                    for part in field.name().split("."):
                        if isinstance(value, dict):
                            value = value.get(part, "")
                        else:
                            value = getattr(value, part, "")
                        if value in (None, ""):
                            break
                else:
                    value = record_dict.get(field.name(), "")
                    if value in (None, "") and hasattr(record, field.name()):
                        # Fallback to attribute access for ActiveRecord instances
                        try:
                            value = getattr(record, field.name())
                        except Exception:
                            value = ""

                row_values.append(field._format_value(value, record_dict))

            ws.append(row_values)

        # --- Custom footer hook ---
        if hasattr(self, "get_export_footer"):
            footer_text = self.get_export_footer()
            if footer_text:
                ws.append([])
                footer_row = ws.max_row + 1
                ws.cell(row=footer_row, column=1, value=footer_text)
                ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(fields))
                cell = ws.cell(row=footer_row, column=1)
                cell.font = Font(italic=True, color="FF666666")
                cell.alignment = Alignment(horizontal="right")

        # --- Auto-fit columns ---
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        download_name = filename or f"{self.__class__.__name__}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
